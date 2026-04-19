"""Extract rare disease data from Excel and produce schema-conformant YAML."""

import re
from pathlib import Path

import click
import openpyxl
import yaml
from oaklib import get_adapter
from oaklib.datamodels.vocabulary import IS_A

# Column mapping: 1-indexed column number -> attribute name
COLUMN_MAP = {
    1: "mondo_id",
    2: "mondo_label",
    3: "mondo_synonyms",
    4: "mondo_categories",
    5: "hpo_high_level_categories",
    6: "histopheno_categories",
    7: "keywords",
    8: "ontology_terminology_codes",
    9: "prevalence_category",
    10: "misdiagnosis_bias",
    11: "prevalence_per_100k_us",
    12: "prioritization_category",
    13: "justification_summary",
    14: "additional_justification",
    15: "hpo_treatment_rank",
    16: "curated_treatments",
    17: "curated_hpo_profiles",
    18: "mondo_category_body_system",
    19: "mondo_category_developmental",
    20: "mondo_category_etiologic",
    21: "mondo_category_genetic",
    22: "mondo_category_extrinsic",
    23: "mondo_category_molecular",
}

HEADER_ROW = 4
DATA_START_ROW = 5

PREVALENCE_MAP = {
    "L": "L",
    "H": "H",
    "H*": "H_star",
    "H?": "H_uncertain",
}

# Fields that should be split into lists
COMMA_SPLIT_FIELDS = {
    "keywords",
    "justification_summary",
}
PIPE_SPLIT_FIELDS: set[str] = set()

MONDO_CATEGORY_FIELDS = (
    "mondo_category_body_system",
    "mondo_category_developmental",
    "mondo_category_etiologic",
    "mondo_category_genetic",
    "mondo_category_extrinsic",
    "mondo_category_molecular",
)

# Fields from OAK (not parsed from Excel)
OAK_FIELDS = {"mondo_label", "mondo_synonyms", "mondo_categories",
               "hpo_high_level_categories", "ontology_terminology_codes"}


def _split_and_strip(value: str, sep: str) -> list[str]:
    """Split a string by separator and strip whitespace from each element."""
    parts = [p.strip().strip('"').strip("'").strip() for p in value.split(sep)]
    return [p for p in parts if p]


def _sanitise_mondo_id(raw: str) -> str | None:
    """Normalise a MONDO ID to the format MONDO:NNNNNNN."""
    if not raw:
        return None
    raw = str(raw).strip()
    match = re.match(r"MONDO[:\s_]*(\d+)", raw)
    if match:
        return f"MONDO:{match.group(1).zfill(7)}"
    return None


def _sanitise_float(value) -> float | None:
    """Convert a value to float, returning None if not possible."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _sanitise_string(value) -> str | None:
    """Convert a cell value to a clean string."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("nr", "n/a", "none", "na", "not found"):
        return None
    return s


class OntologyResolver:
    """Resolves ontology terms using OAK adapters."""

    def __init__(self):
        click.echo("Loading Mondo ontology...")
        self.mondo = get_adapter("sqlite:obo:mondo")
        click.echo("Loading HPO ontology...")
        self.hp = get_adapter("sqlite:obo:hp")
        self._hp_label_cache: dict[str, str | None] = {}
        self._mondo_label_cache: dict[str, str | None] = {}

    def get_mondo_label(self, mondo_id: str) -> str | None:
        return self.mondo.label(mondo_id)

    def get_mondo_synonyms(self, mondo_id: str) -> list[str]:
        label = self.mondo.label(mondo_id)
        aliases = list(self.mondo.entity_aliases(mondo_id))
        # Filter out the label itself
        if label:
            aliases = [a for a in aliases if a.lower() != label.lower()]
        return aliases

    def get_mondo_xrefs(self, mondo_id: str) -> list[str]:
        """Get unique cross-references (xrefs) for a MONDO term."""
        seen = set()
        result = []
        for _pred, obj in self.mondo.simple_mappings_by_curie(mondo_id):
            obj_str = str(obj)
            if obj_str not in seen:
                seen.add(obj_str)
                result.append(obj_str)
        return result

    def get_mondo_categories(self, mondo_id: str) -> list[dict]:
        """Get direct is-a parents as {id, label} dicts."""
        result = []
        for _pred, parent_id in self.mondo.outgoing_relationships(
            mondo_id, predicates=[IS_A]
        ):
            parent_label = self.mondo.label(parent_id)
            if parent_label:
                result.append({"id": parent_id, "label": parent_label})
        return result

    def resolve_hpo_label(self, label: str) -> str | None:
        """Resolve an HPO label to its CURIE."""
        if label in self._hp_label_cache:
            return self._hp_label_cache[label]
        curies = list(self.hp.curies_by_label(label))
        # Filter out blank nodes
        curies = [c for c in curies if not c.startswith("_:")]
        curie = curies[0] if curies else None
        self._hp_label_cache[label] = curie
        return curie

    def resolve_hpo_terms(self, labels: list[str]) -> list[dict]:
        """Resolve HPO term labels to {id, label} dicts."""
        result = []
        for label in labels:
            curie = self.resolve_hpo_label(label.strip())
            if curie:
                result.append({"id": curie, "label": label.strip()})
            else:
                result.append({"id": None, "label": label.strip()})
        return result

    def resolve_mondo_label(self, label: str) -> str | None:
        """Resolve a Mondo term label to its CURIE."""
        if label in self._mondo_label_cache:
            return self._mondo_label_cache[label]
        curies = list(self.mondo.curies_by_label(label))
        curies = [c for c in curies if not c.startswith("_:")]
        curie = curies[0] if curies else None
        self._mondo_label_cache[label] = curie
        return curie

    def resolve_mondo_terms(self, labels: list[str]) -> list[dict]:
        """Resolve Mondo term labels to {id, label} dicts."""
        result = []
        for label in labels:
            curie = self.resolve_mondo_label(label.strip())
            if curie:
                result.append({"id": curie, "label": label.strip()})
            else:
                result.append({"id": None, "label": label.strip()})
        return result


def extract_disease_from_row(row_cells: list) -> dict | None:
    """Extract raw field values from a row of cells. Returns a dict or None."""
    raw = {}
    for col_idx, attr in COLUMN_MAP.items():
        cell = row_cells[col_idx - 1]
        raw[attr] = cell.value

    mondo_id = _sanitise_mondo_id(raw.get("mondo_id"))
    if not mondo_id:
        return None

    raw["mondo_id"] = mondo_id
    return raw


def build_disease(raw: dict, resolver: OntologyResolver) -> dict:
    """Build a disease dict using OAK for label, synonyms, categories, xrefs."""
    mondo_id = raw["mondo_id"]

    # Label from OAK
    mondo_label = resolver.get_mondo_label(mondo_id)
    if not mondo_label:
        return None

    entry: dict = {
        "mondo_id": mondo_id,
        "mondo_label": mondo_label,
    }

    # Synonyms from OAK
    synonyms = resolver.get_mondo_synonyms(mondo_id)
    if synonyms:
        entry["mondo_synonyms"] = synonyms

    # Categories from Excel labels, resolved to {id, label} via OAK
    cat_raw = _sanitise_string(raw.get("mondo_categories"))
    if cat_raw:
        cat_labels = _split_and_strip(cat_raw, ";")
        cat_terms = resolver.resolve_mondo_terms(cat_labels)
        if cat_terms:
            entry["mondo_categories"] = cat_terms

    # HPO categories from Excel labels, resolved via OAK
    hpo_raw = _sanitise_string(raw.get("hpo_high_level_categories"))
    if hpo_raw:
        hpo_labels = _split_and_strip(hpo_raw, ";")
        hpo_terms = resolver.resolve_hpo_terms(hpo_labels)
        if hpo_terms:
            entry["hpo_high_level_categories"] = hpo_terms

    # Xrefs from OAK
    xrefs = resolver.get_mondo_xrefs(mondo_id)
    if xrefs:
        entry["ontology_terminology_codes"] = xrefs

    # Histopheno categories (from Excel, semicolon-separated)
    histo_raw = _sanitise_string(raw.get("histopheno_categories"))
    if histo_raw:
        entry["histopheno_categories"] = _split_and_strip(histo_raw, ";")

    # Keywords (from Excel, comma-separated)
    for field in COMMA_SPLIT_FIELDS:
        val = _sanitise_string(raw.get(field))
        if val:
            entry[field] = _split_and_strip(val, ",")

    # Pipe-split fields
    for field in PIPE_SPLIT_FIELDS:
        val = _sanitise_string(raw.get(field))
        if val:
            entry[field] = _split_and_strip(val, "|")

    # Prevalence category enum
    prev_cat = _sanitise_string(raw.get("prevalence_category"))
    if prev_cat and prev_cat in PREVALENCE_MAP:
        entry["prevalence_category"] = PREVALENCE_MAP[prev_cat]

    # Prioritization category enum
    prio_cat = _sanitise_string(raw.get("prioritization_category"))
    if prio_cat and prio_cat.lower() in ("initial", "expanded"):
        entry["prioritization_category"] = prio_cat.lower()

    # Scalar strings
    for field in (
        "misdiagnosis_bias",
        "additional_justification",
    ):
        val = _sanitise_string(raw.get(field))
        if val:
            entry[field] = val

    # MONDO high-level category fields -> SimpleTerm (semicolon-separated labels)
    for field in MONDO_CATEGORY_FIELDS:
        val = _sanitise_string(raw.get(field))
        if val:
            labels = _split_and_strip(val, ";")
            terms = resolver.resolve_mondo_terms(labels)
            if terms:
                entry[field] = terms

    # Curated HPO profiles -> SimpleTerm (pipe-separated labels)
    hpo_profiles_raw = _sanitise_string(raw.get("curated_hpo_profiles"))
    if hpo_profiles_raw:
        profile_labels = _split_and_strip(hpo_profiles_raw, "|")
        profile_terms = resolver.resolve_hpo_terms(profile_labels)
        if profile_terms:
            entry["curated_hpo_profiles"] = profile_terms

    # Float fields
    for field in ("prevalence_per_100k_us", "hpo_treatment_rank"):
        val = _sanitise_float(raw.get(field))
        if val is not None:
            entry[field] = val

    return entry


def load_treatments(treatments_path: str | Path) -> dict[str, dict]:
    """Load treatments YAML and return a mapping of MONDO ID -> treatment data.

    Each value is a dict with optional 'research' and 'indications' keys.
    """
    path = Path(treatments_path)
    if not path.exists():
        return {}
    with open(path) as f:
        data = yaml.safe_load(f)
    if not data or "diseases" not in data:
        return {}

    result = {}
    for entry in data["diseases"]:
        disease_id = entry.get("disease_id")
        if not disease_id:
            continue
        treatment_data = {}
        if "research" in entry:
            treatment_data["research"] = entry["research"]
        if "indications" in entry:
            treatment_data["indications"] = entry["indications"]
        if treatment_data:
            result[disease_id] = treatment_data
    return result


def extract_all(
    excel_path: str | Path,
    treatments_path: str | Path | None = None,
) -> dict:
    """Extract all diseases from Excel, enrich via OAK, merge treatments."""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # First pass: collect all raw rows and MONDO IDs
    raw_rows = []
    seen_ids = set()
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=23):
        raw = extract_disease_from_row(list(row))
        if raw and raw["mondo_id"] not in seen_ids:
            seen_ids.add(raw["mondo_id"])
            raw_rows.append(raw)
    wb.close()
    click.echo(f"Read {len(raw_rows)} unique diseases from Excel")

    # Load OAK adapters
    resolver = OntologyResolver()

    # Build enriched disease dicts
    diseases = []
    skipped = 0
    for i, raw in enumerate(raw_rows):
        if (i + 1) % 500 == 0:
            click.echo(f"  Processing disease {i + 1}/{len(raw_rows)}...")
        disease = build_disease(raw, resolver)
        if disease:
            diseases.append(disease)
        else:
            skipped += 1

    if skipped:
        click.echo(f"Skipped {skipped} diseases (no label found)")

    # Merge treatments
    treatments = load_treatments(treatments_path) if treatments_path else {}
    if treatments:
        disease_map = {d["mondo_id"]: d for d in diseases}
        merged = 0
        for mondo_id, treatment_data in treatments.items():
            if mondo_id in disease_map:
                if "research" in treatment_data:
                    disease_map[mondo_id]["research"] = treatment_data["research"]
                if "indications" in treatment_data:
                    disease_map[mondo_id]["indications"] = treatment_data["indications"]
                merged += 1
        if merged:
            click.echo(f"Merged treatments for {merged} diseases")

    return {
        "title": "Prioritised Rare Disease List",
        "description": (
            "A list of rare diseases that have been prioritised "
            "for phenotypic characterization research."
        ),
        "version": "0.1",
        "diseases": diseases,
    }


@click.command()
@click.option(
    "--input",
    "-i",
    "input_path",
    required=True,
    type=click.Path(exists=True),
    help="Path to the Excel file.",
)
@click.option(
    "--output",
    "-o",
    "output_path",
    required=True,
    type=click.Path(),
    help="Path to write the output YAML file.",
)
@click.option(
    "--treatments",
    "-t",
    "treatments_path",
    default=None,
    type=click.Path(exists=True),
    help="Path to the treatments YAML file to merge.",
)
def main(input_path: str, output_path: str, treatments_path: str | None):
    """Extract rare disease data from Excel into schema-conformant YAML."""
    data = extract_all(input_path, treatments_path=treatments_path)

    with open(output_path, "w") as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

    click.echo(f"Extracted {len(data['diseases'])} diseases to {output_path}")


if __name__ == "__main__":
    main()
